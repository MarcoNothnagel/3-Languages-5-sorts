import random
import timeit
from openpyxl import load_workbook  # pip install openpyxl
from openpyxl import Workbook


# Timesheets for every sorting algo
bubble = ['Bubble Sort']
merges = ['Merge Sort']
insertion = ['Insertion Sort']
shell = ['Shell Sort']
selection = ['Selection Sort']


# Time Complexity: O(n^2)
def bubble_sort(array):
    for i in range(len(array) - 1, 0, -1):
        for j in range(i):
            if array[j] > array[j + 1]:
                temp = array[j]
                array[j] = array[j + 1]
                array[j + 1] = temp


# Time Complexity: O(n log(n))
def merge_sort(array):
    if len(array) <= 1:
        return array
    middle = len(array) // 2
    left = array[:middle]
    right = array[middle:]

    left = merge_sort(left)
    right = merge_sort(right)
    return list(merge(left, right))


# Merge the sorted halves
def merge(left, right):
    res = []
    while len(left) != 0 and len(right) != 0:
        if left[0] < right[0]:
            res.append(left[0])
            left.remove(left[0])
        else:
            res.append(right[0])
            right.remove(right[0])
    if len(left) == 0:
        res = res + right
    else:
        res = res + left
    return res


# Time Complexity: O(n^2)
def insertion_sort(array):
    for i in range(1, len(array)):
        a = array[i]
        j = i - 1
        while j >= 0 and a < array[j]:
            array[j + 1] = array[j]
            j -= 1
        array[j + 1] = a
    return array


# Time Complexity: O(n^2)
def shell_sort(array):
    gap = len(array) // 2
    while gap > 0:
        for i in range(gap, len(array)):
            temp = array[i]
            j = i
            while j >= gap and array[j - gap] > temp:
                array[j] = array[j - gap]
                j -= gap
            array[j] = temp
        gap //= 2


# Time Complexity: O(n^2)
def selection_sort(array):
    for i in range(len(array)):
        j = i
        for k in range(i + 1, len(array)):
            # For sorting in descending order
            # for minimum element in each loop
            if array[k] < array[j]:
                j = k
        # Arranging min at the correct position
        (array[i], array[j]) = (array[j], array[i])


class color:
    BOLD = '\033[1m'
    END = '\033[0m'


class heading:
    UNDERLINE = '\033[4m'
    END = '\033[0m'


print(heading.UNDERLINE + 'Sorting 100 items' + heading.END, '\n')
print(color.BOLD + 'Time in milli seconds')
unsorted_list = random.sample(range(0, 100), 100)


print(color.BOLD + 'Bubble Sort' + color.END, end='' '\t' '\t')
start_time = timeit.default_timer()
bubble_sort(unsorted_list)
elapsed = (timeit.default_timer() - start_time) * pow(10, 3)
bubble.append(elapsed)
print('Time elapsed:', elapsed)


print(color.BOLD + 'Merge Sort' + color.END, end='' '\t' '\t')
start_time = timeit.default_timer()
merge_sort(unsorted_list)
elapsed = (timeit.default_timer() - start_time) * pow(10, 3)
merges.append(elapsed)
print('Time elapsed:', elapsed)


print(color.BOLD + 'Insertion Sort' + color.END, end='' '\t' '\t')
start_time = timeit.default_timer()
insertion_sort(unsorted_list)
elapsed = (timeit.default_timer() - start_time) * pow(10, 3)
insertion.append(elapsed)
print('Time elapsed:', elapsed)

print(color.BOLD + 'Shell Sort' + color.END, end='' '\t' '\t')
start_time = timeit.default_timer()
shell_sort(unsorted_list)
elapsed = (timeit.default_timer() - start_time) * pow(10, 3)
shell.append(elapsed)
print('Time elapsed:', elapsed)

print(color.BOLD + 'Selection Sort' + color.END, end='' '\t' '\t')
start_time = timeit.default_timer()
selection_sort(unsorted_list)
elapsed = (timeit.default_timer() - start_time) * pow(10, 3)
selection.append(elapsed)
print('Time elapsed:', elapsed)


print('\n', heading.UNDERLINE + 'Sorting 1000 items' + heading.END, '\n')
print(color.BOLD + 'Time in milli seconds')
unsorted_list = random.sample(range(0, 1000), 1000)


print(color.BOLD + 'Bubble Sort' + color.END, end='' '\t' '\t')
start_time = timeit.default_timer()
bubble_sort(unsorted_list)
elapsed = (timeit.default_timer() - start_time) * pow(10, 3)
bubble.append(elapsed)
print('Time elapsed:', elapsed)


print(color.BOLD + 'Merge Sort' + color.END, end='' '\t' '\t')
start_time = timeit.default_timer()
merge_sort(unsorted_list)
elapsed = (timeit.default_timer() - start_time) * pow(10, 3)
merges.append(elapsed)
print('Time elapsed:', elapsed)


print(color.BOLD + 'Insertion Sort' + color.END, end='' '\t' '\t')
start_time = timeit.default_timer()
insertion_sort(unsorted_list)
elapsed = (timeit.default_timer() - start_time) * pow(10, 3)
insertion.append(elapsed)
print('Time elapsed:', elapsed)

print(color.BOLD + 'Shell Sort' + color.END, end='' '\t' '\t')
start_time = timeit.default_timer()
shell_sort(unsorted_list)
elapsed = (timeit.default_timer() - start_time) * pow(10, 3)
shell.append(elapsed)
print('Time elapsed:', elapsed)

print(color.BOLD + 'Selection Sort' + color.END, end='' '\t' '\t')
start_time = timeit.default_timer()
selection_sort(unsorted_list)
elapsed = (timeit.default_timer() - start_time) * pow(10, 3)
selection.append(elapsed)
print('Time elapsed:', elapsed)


print('\n', heading.UNDERLINE + 'Sorting 10000 items' + heading.END, '\n')
print(color.BOLD + 'Time in milli seconds')
unsorted_list = random.sample(range(0, 10000), 10000)

print(color.BOLD + 'Bubble Sort' + color.END, end='' '\t' '\t')
start_time = timeit.default_timer()
bubble_sort(unsorted_list)
elapsed = (timeit.default_timer() - start_time) * pow(10, 3)
bubble.append(elapsed)
print('Time elapsed:', elapsed)

print(color.BOLD + 'Merge Sort' + color.END, end='' '\t' '\t')
start_time = timeit.default_timer()
merge_sort(unsorted_list)
elapsed = (timeit.default_timer() - start_time) * pow(10, 3)
merges.append(elapsed)
print('Time elapsed:', elapsed)

print(color.BOLD + 'Insertion Sort' + color.END, end='' '\t' '\t')
start_time = timeit.default_timer()
insertion_sort(unsorted_list)
elapsed = (timeit.default_timer() - start_time) * pow(10, 3)
insertion.append(elapsed)
print('Time elapsed:', elapsed)

print(color.BOLD + 'Shell Sort' + color.END, end='' '\t' '\t')
start_time = timeit.default_timer()
shell_sort(unsorted_list)
elapsed = (timeit.default_timer() - start_time) * pow(10, 3)
shell.append(elapsed)
print('Time elapsed:', elapsed)

print(color.BOLD + 'Selection Sort' + color.END, end='' '\t' '\t')
start_time = timeit.default_timer()
selection_sort(unsorted_list)
elapsed = (timeit.default_timer() - start_time) * pow(10, 3)
selection.append(elapsed)
print('Time elapsed:', elapsed)


header = ['Sorting method', 'Time elapsed 100',
          'Time elapsed 1000', 'Time elapsed 10000']


workbook = load_workbook(
    filename="C:/Users/Marco/OneDrive/BSc IT/Y3B3/python/export.xlsx")

worksheet = workbook.active

row = 2
for item in header:
    worksheet.cell(row=row, column=(header.index(item) + 1), value=item)

row += 1
for item in bubble:
    worksheet.cell(row=row, column=(bubble.index(item) + 1), value=item)

row += 1
for item in merges:
    worksheet.cell(row=row, column=(merges.index(item) + 1), value=item)

row += 1
for item in insertion:
    worksheet.cell(row=row, column=(insertion.index(item) + 1), value=item)

row += 1
for item in shell:
    worksheet.cell(row=row, column=(shell.index(item) + 1), value=item)

row += 1
for item in selection:
    worksheet.cell(row=row, column=(selection.index(item) + 1), value=item)


file = open("javafile.txt", "r")

bubble.clear()
line = file.readline()
bubble = line.split('/')
bubble = ['Bubble Sort', float(bubble[1]), float(bubble[2]), float(bubble[3])]

merges.clear()
line = file.readline()
merges = line.split('/')
merges = ['Merge Sort', float(merges[1]), float(merges[2]), float(merges[3])]

insertion.clear()
line = file.readline()
insertion = line.split('/')
insertion = ['Insertion Sort', float(insertion[1]), float(
    insertion[2]), float(insertion[3])]

shell.clear()
line = file.readline()
shell = line.split('/')
shell = ['Shell Sort', float(shell[1]), float(shell[2]), float(shell[3])]

selection.clear()
line = file.readline()
selection = line.split('/')
selection = ['Selection Sort', float(selection[1]), float(
    selection[2]), float(selection[3])]

file.close()

row = 36
for item in header:
    worksheet.cell(row=row, column=(header.index(item) + 1), value=item)

row += 1
for item in bubble:
    worksheet.cell(row=row, column=(bubble.index(item) + 1), value=item)

row += 1
for item in merges:
    worksheet.cell(row=row, column=(merges.index(item) + 1), value=item)

row += 1
for item in insertion:
    worksheet.cell(row=row, column=(insertion.index(item) + 1), value=item)

row += 1
for item in shell:
    worksheet.cell(row=row, column=(shell.index(item) + 1), value=item)

row += 1
for item in selection:
    worksheet.cell(row=row, column=(selection.index(item) + 1), value=item)


file = open("cppfile.txt", "r")

bubble.clear()
line = file.readline()
bubble = line.split('/')
bubble = ['Bubble Sort', float(bubble[1]), float(bubble[2]), float(bubble[3])]

merges.clear()
line = file.readline()
merges = line.split('/')
merges = ['Merge Sort', float(merges[1]), float(merges[2]), float(merges[3])]

insertion.clear()
line = file.readline()
insertion = line.split('/')
insertion = ['Insertion Sort', float(insertion[1]), float(
    insertion[2]), float(insertion[3])]

shell.clear()
line = file.readline()
shell = line.split('/')
shell = ['Shell Sort', float(shell[1]), float(shell[2]), float(shell[3])]

selection.clear()
line = file.readline()
selection = line.split('/')
selection = ['Selection Sort', float(selection[1]), float(
    selection[2]), float(selection[3])]

file.close()

row = 66
for item in header:
    worksheet.cell(row=row, column=(header.index(item) + 1), value=item)

row += 1
for item in bubble:
    worksheet.cell(row=row, column=(bubble.index(item) + 1), value=item)

row += 1
for item in merges:
    worksheet.cell(row=row, column=(merges.index(item) + 1), value=item)

row += 1
for item in insertion:
    worksheet.cell(row=row, column=(insertion.index(item) + 1), value=item)

row += 1
for item in shell:
    worksheet.cell(row=row, column=(shell.index(item) + 1), value=item)

row += 1
for item in selection:
    worksheet.cell(row=row, column=(selection.index(item) + 1), value=item)

workbook.save(
    filename="C:/Users/Marco/OneDrive/BSc IT/Y3B3/python/export.xlsx")
